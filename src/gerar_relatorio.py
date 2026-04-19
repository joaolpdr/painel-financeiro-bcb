import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ── Estilos ───────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=11, color="2c3e50"):
    return Font(bold=True, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def borda():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

AZUL   = "2980b9"
PRETO  = "2c3e50"
CINZA  = "f2f2f2"
BRANCO = "FFFFFF"
VERDE  = "2ecc71"
AMARELO= "f39c12"
VERMELHO = "e74c3c"


# ── Helpers ───────────────────────────────────────────────
def cabecalho_aba(ws, titulo):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws["A1"].font = bold(13, BRANCO)
    ws["A1"].fill = fill(AZUL)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Fonte: Banco Central do Brasil  |  Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

def linha_header(ws, row, colunas):
    for col_idx, nome in enumerate(colunas, start=1):
        cell = ws.cell(row=row, column=col_idx, value=nome)
        cell.font = bold(10, BRANCO)
        cell.fill = fill(PRETO)
        cell.alignment = center()
        cell.border = borda()
    ws.row_dimensions[row].height = 22

def linha_dados(ws, row, valores, cor=CINZA):
    for col_idx, valor in enumerate(valores, start=1):
        cell = ws.cell(row=row, column=col_idx, value=valor)
        cell.fill = fill(cor)
        cell.alignment = center()
        cell.border = borda()
    ws.row_dimensions[row].height = 20


# ════════════════════════════════════════════════════════
# ABA 1 — Resumo Executivo
# ════════════════════════════════════════════════════════
def aba_resumo(wb, df_ptax, df_series):
    ws = wb.create_sheet("Resumo Executivo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28

    ws.merge_cells("A1:B1")
    ws["A1"] = "PAINEL FINANCEIRO — BANCO CENTRAL DO BRASIL"
    ws["A1"].font = bold(14, BRANCO)
    ws["A1"].fill = fill(AZUL)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Dados reais coletados via API  |  Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # Cabeçalho tabela
    for col, titulo in [("A", "Indicador"), ("B", "Valor")]:
        ws[f"{col}4"].value = titulo
        ws[f"{col}4"].font = bold(11, BRANCO)
        ws[f"{col}4"].fill = fill(PRETO)
        ws[f"{col}4"].alignment = center()
        ws[f"{col}4"].border = borda()
    ws.row_dimensions[4].height = 24

    # Métricas por moeda
    cores = {"Dólar": "e8f4fc", "Euro": "e8f8f0", "Libra": "fde8e8"}
    metricas = []

    for moeda in ["Dólar", "Euro", "Libra"]:
        grupo = df_ptax[df_ptax["moeda"] == moeda].copy()
        if not grupo.empty:
            ultimo  = grupo.iloc[-1]["venda"]
            primeiro = grupo.iloc[0]["venda"]
            variacao = round(((ultimo - primeiro) / primeiro) * 100, 2)
            sinal = "▲" if variacao > 0 else "▼"
            metricas.append((
                f"Cotação Atual — {moeda}",
                f"R$ {ultimo:.4f}  {sinal} {abs(variacao)}% (12m)",
                cores[moeda]
            ))

    # Métricas de juros
    for serie, cor in [("Selic", "fff8e8"), ("CDI", "fff0e8"), ("IPCA", "fde8e8")]:
        grupo = df_series[df_series["serie"] == serie]
        if not grupo.empty:
            ultimo_valor = grupo.iloc[-1]["valor"]
            metricas.append((
                f"Última Taxa — {serie}",
                f"{ultimo_valor:.2f}% a.m.",
                cor
            ))

    # Período dos dados
    df_ptax["data"] = pd.to_datetime(df_ptax["data"])
    metricas.append((
        "Período dos Dados",
        f"{df_ptax['data'].min().strftime('%d/%m/%Y')} a {df_ptax['data'].max().strftime('%d/%m/%Y')}",
        CINZA
    ))

    metricas.append((
        "Total de Registros Coletados",
        f"{len(df_ptax) + len(df_series)} registros",
        CINZA
    ))

    for i, (indicador, valor, cor) in enumerate(metricas, start=5):
        ws[f"A{i}"] = indicador
        ws[f"B{i}"] = valor
        for col in ["A", "B"]:
            ws[f"{col}{i}"].fill = fill(cor)
            ws[f"{col}{i}"].alignment = center()
            ws[f"{col}{i}"].border = borda()
        ws.row_dimensions[i].height = 22


# ════════════════════════════════════════════════════════
# ABA 2 — Cotações
# ════════════════════════════════════════════════════════
def aba_cotacoes(wb, df_ptax):
    ws = wb.create_sheet("Cotações")
    cabecalho_aba(ws, "COTAÇÕES PTAX — BANCO CENTRAL DO BRASIL")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    linha_header(ws, 4, ["Data", "Moeda", "Compra (R$)", "Venda (R$)"])

    cores_moeda = {"Dólar": "e8f4fc", "Euro": "e8f8f0", "Libra": "fde8e8"}

    for i, row in df_ptax.iterrows():
        cor = cores_moeda.get(row["moeda"], CINZA)
        linha_dados(ws, i + 5, [
            str(row["data"]),
            row["moeda"],
            f"R$ {float(row['compra']):.4f}",
            f"R$ {float(row['venda']):.4f}",
        ], cor)


# ════════════════════════════════════════════════════════
# ABA 3 — Séries Econômicas
# ════════════════════════════════════════════════════════
def aba_series(wb, df_series):
    ws = wb.create_sheet("Séries Econômicas")
    cabecalho_aba(ws, "SÉRIES ECONÔMICAS — SELIC, CDI E IPCA")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    linha_header(ws, 4, ["Data", "Série", "Valor (%)"])

    cores_serie = {"Selic": "e8f4fc", "CDI": "fde8e8", "IPCA": "fff8e8"}

    for i, row in df_series.iterrows():
        cor = cores_serie.get(row["serie"], CINZA)
        linha_dados(ws, i + 5, [
            str(row["data"]),
            row["serie"],
            f"{float(row['valor']):.4f}%",
        ], cor)


# ════════════════════════════════════════════════════════
# ABA 4 — Gráficos
# ════════════════════════════════════════════════════════
def aba_graficos(wb):
    ws = wb.create_sheet("Gráficos")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    ws["A1"] = "PAINÉIS VISUAIS — PAINEL FINANCEIRO BCB"
    ws["A1"].font = bold(13, BRANCO)
    ws["A1"].fill = fill(AZUL)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    graficos = [
        ("output/01_evolucao_cotacoes.png", "A3"),
        ("output/02_variacao_mensal.png",   "H3"),
        ("output/03_selic_cdi.png",         "A22"),
        ("output/04_ipca.png",              "H22"),
    ]

    for caminho, ancora in graficos:
        if os.path.exists(caminho):
            img        = Image(caminho)
            img.width  = 480
            img.height = 280
            ws.add_image(img, ancora)


# ── Execução principal ────────────────────────────────────
if __name__ == "__main__":
    df_ptax   = pd.read_csv("data/raw/cotacoes.csv")
    df_series = pd.read_csv("data/raw/series_economicas.csv")

    wb = Workbook()
    wb.remove(wb.active)

    aba_resumo(wb,   df_ptax, df_series)
    aba_cotacoes(wb, df_ptax)
    aba_series(wb,   df_series)
    aba_graficos(wb)

    output = "output/painel_financeiro_bcb.xlsx"
    wb.save(output)
    print(f"✅ Relatório gerado: {output}")